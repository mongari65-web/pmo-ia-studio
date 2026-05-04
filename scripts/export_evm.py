#!/usr/bin/env python3
"""
export_evm.py — Generate EVM Excel with S-curve chart
Usage: python3 export_evm.py input.json output.xlsx
"""
import sys, json, io, base64
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_B64 = "UEsDBBQAAAAIACR4pFxGx01IlQAAAM0AAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE3PTQvCMAwG4L9SdreZih6kDkQ9ip68zy51hbYpbYT67+0EP255ecgboi6JIia2mEXxLuRtMzLHDUDWI/o+y8qhiqHke64x3YGMsRoPpB8eA8OibdeAhTEMOMzit7Dp1C5GZ3XPlkJ3sjpRJsPiWDQ6sScfq9wcChDneiU+ixNLOZcrBf+LU8sVU57mym/8ZAW/B7oXUEsDBBQAAAAIACR4pFwWCiOq7wAAACsCAAARAAAAZG9jUHJvcHMvY29yZS54bWzNks9qwzAMh19l+J7ISdfCTJrLRk8dDFbY2M3Yamsa/8HWSPr2c7I2ZWwPsKOlnz99AjUqCOUjvkQfMJLBdDfYziWhwpodiYIASOqIVqYyJ1xu7n20kvIzHiBIdZIHhJrzFVgkqSVJGIFFmImsbbQSKqIkHy94rWZ8+IzdBNMKsEOLjhJUZQWsHSeG89A1cAOMMMJo03cB9Uycqn9ipw6wS3JIZk71fV/2iymXd6jg/Xn7Oq1bGJdIOoX5VzKCzgHX7Dr5bfH4tNuwtub1quDLgt/vqqXgleAPH6PrD7+bsPXa7M0/Nr4Ktg38uov2C1BLAwQUAAAACAAkeKRcmVycIxAGAACcJwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54bWztWltz2jgUfu+v0Hhn9m0LxjaBtrQTc2l227SZhO1OH4URWI1seWSRhH+/RzYQy5YN7ZJNups8BCzp+85FR+foOHnz7i5i6IaIlPJ4YNkv29a7ty/e4FcyJBFBMBmnr/DACqVMXrVaaQDDOH3JExLD3IKLCEt4FMvWXOBbGi8j1uq0291WhGlsoRhHZGB9XixoQNBUUVpvXyC05R8z+BXLVI1lowETV0EmuYi08vlsxfza3j5lz+k6HTKBbjAbWCB/zm+n5E5aiOFUwsTAamc/VmvH0dJIgILJfZQFukn2o9MVCDINOzqdWM52fPbE7Z+Mytp0NG0a4OPxeDi2y9KLcBwE4FG7nsKd9Gy/pEEJtKNp0GTY9tqukaaqjVNP0/d93+ubaJwKjVtP02t33dOOicat0HgNvvFPh8Ouicar0HTraSYn/a5rpOkWaEJG4+t6EhW15UDTIABYcHbWzNIDll4p+nWUGtkdu91BXPBY7jmJEf7GxQTWadIZljRGcp2QBQ4AN8TRTFB8r0G2iuDCktJckNbPKbVQGgiayIH1R4Ihxdyv/fWXu8mkM3qdfTrOa5R/aasBp+27m8+T/HPo5J+nk9dNQs5wvCwJ8fsjW2GHJ247E3I6HGdCfM/29pGlJTLP7/kK6048Zx9WlrBdz8/knoxyI7vd9lh99k9HbiPXqcCzIteURiRFn8gtuuQROLVJDTITPwidhphqUBwCpAkxlqGG+LTGrBHgE323vgjI342I96tvmj1XoVhJ2oT4EEYa4pxz5nPRbPsHpUbR9lW83KOXWBUBlxjfNKo1LMXWeJXA8a2cPB0TEs2UCwZBhpckJhKpOX5NSBP+K6Xa/pzTQPCULyT6SpGPabMjp3QmzegzGsFGrxt1h2jSPHr+BfmcNQockRsdAmcbs0YhhGm78B6vJI6arcIRK0I+Yhk2GnK1FoG2camEYFoSxtF4TtK0EfxZrDWTPmDI7M2Rdc7WkQ4Rkl43Qj5izouQEb8ehjhKmu2icVgE/Z5ew0nB6ILLZv24fobVM2wsjvdH1BdK5A8mpz/pMjQHo5pZCb2EVmqfqoc0PqgeMgoF8bkePuV6eAo3lsa8UK6CewH/0do3wqv4gsA5fy59z6XvufQ9odK3NyN9Z8HTi1veRm5bxPuuMdrXNC4oY1dyzcjHVK+TKdg5n8Ds/Wg+nvHt+tkkhK+aWS0jFpBLgbNBJLj8i8rwKsQJ6GRbJQnLVNNlN4oSnkIbbulT9UqV1+WvuSi4PFvk6a+hdD4sz/k8X+e0zQszQ7dyS+q2lL61JjhK9LHMcE4eyww7ZzySHbZ3oB01+/ZdduQjpTBTl0O4GkK+A226ndw6OJ6YkbkK01KQb8P56cV4GuI52QS5fZhXbefY0dH758FRsKPvPJYdx4jyoiHuoYaYz8NDh3l7X5hnlcZQNBRtbKwkLEa3YLjX8SwU4GRgLaAHg69RAvJSVWAxW8YDK5CifEyMRehw55dcX+PRkuPbpmW1bq8pdxltIlI5wmmYE2eryt5lscFVHc9VW/Kwvmo9tBVOz/5ZrcifDBFOFgsSSGOUF6ZKovMZU77nK0nEVTi/RTO2EpcYvOPmx3FOU7gSdrYPAjK5uzmpemUxZ6by3y0MCSxbiFkS4k1d7dXnm5yueiJ2+pd3wWDy/XDJRw/lO+df9F1Drn723eP6bpM7SEycecURAXRFAiOVHAYWFzLkUO6SkAYTAc2UyUTwAoJkphyAmPoLvfIMuSkVzq0+OX9FLIOGTl7SJRIUirAMBSEXcuPv75Nqd4zX+iyBbYRUMmTVF8pDicE9M3JD2FQl867aJguF2+JUzbsaviZgS8N6bp0tJ//bXtQ9tBc9RvOjmeAes4dzm3q4wkWs/1jWHvky3zlw2zreA17mEyxDpH7BfYqKgBGrYr66r0/5JZw7tHvxgSCb/NbbpPbd4Ax81KtapWQrET9LB3wfkgZjjFv0NF+PFGKtprGtxtoxDHmAWPMMoWY434dFmhoz1YusOY0Kb0HVQOU/29QNaPYNNByRBV4xmbY2o+ROCjzc/u8NsMLEjuHti78BUEsDBBQAAAAIACR4pFyVniUOEwEAAMwBAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sTVFdT8MgFP0rhB8wOpOpWdom24zRB5NmRn1m621LBtwKt1b/vUDXZk+ccz8O50A+orv4DoDYr9HWF7wj6rdC+HMHRvoV9mBDp0FnJAXqWuF7B7JOS0aLuyy7F0Yqy8s81SpX5jiQVhYqx/xgjHR/e9A4FnzN58JRtR3FgijzXrbwDvTRVy4wsajUyoD1Ci1z0BR8t97u0nwa+FQw+hvMYpIT4iWS17rgWTQEGs4UFWQ4fuAAWkehYOP7qsmXK+PiLZ7Vn1P2kOUkPRxQf6mauoI/clZDIwdNRxxf4Jpnsxh8kiRnuQnHnG/Stcp6pqEJ49nqYcOZm3YnQtindzohEZoEu/Dc4OJA6DeINJNoffnA8h9QSwMEFAAAAAgAJHikXJWeJQ4TAQAAzAEAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWxNUV1PwyAU/SuEHzA6k6lZ2ibbjNEHk2ZGfWbrbUsG3Aq3Vv+9QNdmT5xzPw7nQD6iu/gOgNiv0dYXvCPqt0L4cwdG+hX2YEOnQWckBepa4XsHsk5LRou7LLsXRirLyzzVKlfmOJBWFirH/GCMdH970DgWfM3nwlG1HcWCKPNetvAO9NFXLjCxqNTKgPUKLXPQFHy33u7SfBr4VDD6G8xikhPiJZLXuuBZNAQazhQVZDh+4ABaR6Fg4/uqyZcr4+ItntWfU/aQ5SQ9HFB/qZq6gj9yVkMjB01HHF/gmmezGHySJGe5Ccecb9K1ynqmoQnj2ephw5mbdidC2Kd3OiERmgS78Nzg4kDoN4g0k2h9+cDyH1BLAwQUAAAACAAkeKRclZ4lDhMBAADMAQAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQzLnhtbE1RXU/DIBT9K4QfMDqTqVnaJtuM0QeTZkZ9ZuttSwbcCrdW/71A12ZPnHM/DudAPqK7+A6A2K/R1he8I+q3QvhzB0b6FfZgQ6dBZyQF6lrheweyTktGi7ssuxdGKsvLPNUqV+Y4kFYWKsf8YIx0f3vQOBZ8zefCUbUdxYIo81628A700VcuMLGo1MqA9Qotc9AUfLfe7tJ8GvhUMPobzGKSE+Ilkte64Fk0BBrOFBVkOH7gAFpHoWDj+6rJlyvj4i2e1Z9T9pDlJD0cUH+pmrqCP3JWQyMHTUccX+CaZ7MYfJIkZ7kJx5xv0rXKeqahCePZ6mHDmZt2J0LYp3c6IRGaBLvw3ODiQOg3iDSTaH35wPIfUEsDBBQAAAAIACR4pFyVniUOEwEAAMwBAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1sTVFdT8MgFP0rhB8wOpOpWdom24zRB5NmRn1m621LBtwKt1b/vUDXZk+ccz8O50A+orv4DoDYr9HWF7wj6rdC+HMHRvoV9mBDp0FnJAXqWuF7B7JOS0aLuyy7F0Yqy8s81SpX5jiQVhYqx/xgjHR/e9A4FnzN58JRtR3FgijzXrbwDvTRVy4wsajUyoD1Ci1z0BR8t97u0nwa+FQw+hvMYpIT4iWS17rgWTQEGs4UFWQ4fuAAWkehYOP7qsmXK+PiLZ7Vn1P2kOUkPRxQf6mauoI/clZDIwdNRxxf4Jpnsxh8kiRnuQnHnG/Stcp6pqEJ49nqYcOZm3YnQtindzohEZoEu/Dc4OJA6DeINJNoffnA8h9QSwMEFAAAAAgAJHikXAUC4WwZBAAAIhYAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0NS54bWyNmO1uozgUhm8FcQFD7Hy2SiK1sWen0oyEptrZ3zSYxFPArHGazl79GkwhlnyO+qcFHh/78SHKq3h7Vfq1PQthoveqrNtdfDamuU+S9ngWVdZ+UY2oLSmUrjJjb/UpaRstsrwvqsqEzmarpMpkHe+3/bNU77fqYkpZi1RH7aWqMv3nUZTquotJ/PHgpzydTfcg2W+b7CSehfm7SbW9S8ZZclmJupWqjrQodvEDved01RX0I35JcW1vrqP2rK5/aZl/tyvbjcziqNvci1KvHX7Ku0fdYrWI3p+bUvbLR3+GSxpHRjXfRWEOoix38eM8jrKjkW8itRW7+EUZo6pe227CZMY+K7T6T9S9kSiFHWxVm360ncoNDTA3U7cSTId1nIATeuia96/rRN+2ZNz67fVHS772b8y+gZesFQdV/iNzc97FmzjKRZFdSvNTXb+J4S0su/mOqmz7v9HVjSW2J8dLa3WGYmtQydr9z96Ht3dbsAAK6FBAP1swHwrmny1YDAWLzxYsh4J+64nbe984lplsv9XqGulevGsQ3XzMMrbMfpSO3YgHO6Ttp7VPZd197J+NtlTaCc0+FVqqXGwTY1fpHiXHofDRFVKo8Fd0uFSXMjOyCFQfXPUcqOZ4NXPVC6D64YBWc1e9hNa25alWv93H2q9PbF/H5s7HHkLbSEmocW5094Xztp9tk7fbtiCMIYyHmae7GHWhvqU0pLtAdBHGEMbDzNNdjrrQi0rnId0looswhjAeZp7uatRdQbqLkO4K0UUYQxgPM093PequId1lSHeN6CKMIYyHmae7GXU3kO4qpLtBdBHGEMbDzNO9G3XvIN11SPcO0UUYQxgPM0+XzEZfewkIb0LCw/CwMQYZBjkAfekpwgiYXndBaYJJI5BhkAPQl6aTNJicZBa0ppg1AhkGOQB96ynpCBx1wawjWNhhkGGQA9C3ngKPgIlHgpFHsMzDIMMgB6BvPeUeAYOPBJOPYNGHQYZBDkDfeoo/AuYfCQYgwRIQgwyDHIC+9ZSCBIxBEsxBggUhBhkGOQB96ykMCZiGJBiHBMtDDDIMcgD61lMmEjAUSTAVCRaLGGQY5AD0rOmUjBRMRhKMRopFIwYZBjkAfeubX3dgNJJgNlIsGzHIMMgB6FtP2UjBbKTBbKRYNmKQYZAD0LeespGC2UiD2UixbMQgwyAHoG89ZSOFfw0Gs5Fi2YhBhkEOQN96ykYKZiMNZiPFshGDDIMcgL71lI0UzEYazEaKZSMGGQY5AJ11cnOK1J1s/sj0SdZtVIrC1sy+rO2WtTtEcjdGNf35kTv56y/PIsuF7gZYXihlPm6S/TbX2VXWJ3dee68/c2KrikIeBVPHSyVq445steiOd1TdnmXTWqN7me9i/ZS7o8XxWHj/P1BLAwQUAAAACAAkeKRc5adt3icBAADBAgAAGAAAAHhsL2RyYXdpbmdzL2RyYXdpbmcxLnhtbJ2S3U7EIBCFX4XwANI1Wk1DScw2Gm/UV5hQWEj4aQZ2t7690PZiV2/M9oZzOJ0PGODnNCCZvQuppybnqWMsSaM8pLs4qVASHdFDLhYPbEQ423Dwjt03TcvShArGZJTKw5pQwWNQe+XcS5AmouAaoxdcRidazupQ9afWollsVRzjWew4q0PVW7opzlaEmjORc0+f24emfJTI754+PrWLYYIfECZj5SuCV4KH09uF/yobkR+nLyR27OmOklAme7o3gJnsanVJrwvKHPsLmTX6kkC3rba2roMbmufBBnqBGiADOaK9ASXrOQpLdovaNiVvJm0A/A8gam2lGqI8ehXySkHlINsYkrFTogS72nV8H5dOs6sTX/qiry9ROluY9bda9+thsfp0xQ9QSwMEFAAAAAgAJHikXI7R6TmLAAAA7QAAACMAAAB4bC9kcmF3aW5ncy9fcmVscy9kcmF3aW5nMS54bWwucmVsc43POwrDMBAE0KuIPYDXTpEiyK7SuA2+wCKvLRHrg7QB5/YRJIUDKVLODDwYfeONxMVQrEtF7X4LpQcrki6IxVj2VJqYONRlidmT1JhXTGTutDKe2vaM+WjAoI+mmp6J/xHjsjjD12genoP8gNFYygJqoryy9ID79q4+S9dUEdQ495DHuQMcNH69G15QSwMEFAAAAAgAJHikXHf3IoWNAAAA8wAAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0NS54bWwucmVsc43POwrDMBAE0KuIPYDXTpEiyK7SuA2+wCKvPsT6IMnEuX3cJDiQIuXMwIORN16ouhiKdamIzS+h9GBrTRfEoix7Kk1MHPZFx+yp7jEbTKTuZBhPbXvGfDRgkEdTTM/E/4hRa6f4GtXqOdQfMM6ZHi4YEBNlw7UH3JZ3+Vm7ZndBjHMPeZw7wEHi18fhBVBLAwQUAAAACAAkeKRcQ7mUqgADAACOEQAAFAAAAHhsL2NoYXJ0cy9jaGFydDEueG1s3Zj/btowEMdfJUsrdZOmOQVBVRQiQaDTpKGiduv/JjbBq2OntlnD/tpD7An3JPOPwICKNR0gUP9J6tz57Pve55CuYTKBQt3mMMFekVEm2/5EqbwFgEwmOIPyA88x05YxFxlUeilSgAR8JCzNKKgFQRPYEH4USjWj2PsOads/D3wQhdYQhYooivWriEJBkkkUwtaIo9lQuBNb8D/OzCBh5gjYyreNYoMMhXkhPL4ZCh0WLD7Zh4piPhUj7N16v3/+8vpQMIy8O0in2BtABlOcYabMLmX3ChdBP13CwOQOSh1yylVHYBiFlDAcO41Swae5vpmTTyrIEBTIZCixjkZQ4SxWVy4QFktrE10qcYPHUTiOzhZ3xcobYCXIwxTLszfdWgi0A5h72kvJMk3KtpXRe2z7tUYQBFZQySlBV4RSuxDpKKbljWuNZr3f9Z3IK265kKoH5aTUwJhKN8rMve1VMyjujSJylo04da4JEQnFVizyoySwYZdldotjdsFKxWy2V3SvOoK5kAnU+LFp9gw9p53Teks/ak1H0XwHsPv1SZWCdE2Q7tMgdj+wqK/yfr7G+3l13uPj4L13eXERNHfKOyIw4wwdBPjN6ewb+C2FPAzwsQkSvwD42hrwterA944D+KBx2Wxe7hR4nSJk6YF+4jfns2/it1TyMMT3TJDeC4ivrxFfr058/ziIv4jrnX7v+Toh/Vc14OXDFIrD4L45mX3jXlXGo8K9b4L0/4k7LD6h5bFoeW0/gKVBQB/eKZ7ukQnUTqlpFKJHDagIZ86eETaAhQ2zcILFkEtnprabXsEENsSCcIRlxSkrg9+4+EKS+4GmwknBOLMtpRXbZEoEl7JTrFSHjuj1eCx13VdqVhZKf1uv1w4KZm//URBkyJCvpIK6hZT3tv/15l3FGupmusqU5w6PdfHb/sn7k5PA96TuxgR/Juweo3II3knBbVXKkoKlQR2nmKH5e1Eq4Zq3NBr3OyKvGZ0tzSqIyLxLIbuXnXJXCnMHkGt48Pc/MNEfUEsDBBQAAAAIACR4pFyoxusLJQMAAE0QAAANAAAAeGwvc3R5bGVzLnhtbN1YUW/aMBD+K1F+wBLICGQCpJK20qRtqtQ+7NUQJ7HkxJljOuivn88OJICvYlsf1gWh2Hf+vvvOOTuGeav2nD6WlCpvV/G6XfilUs2nIGg3Ja1I+0E0tNaeXMiKKN2VRdA2kpKsBVDFg3EYxkFFWO0v5/W2uq9U623EtlYLP/SD5TwXdW8Z+9agh5KKes+EL/yUcLaWzIwlFeN7ax6DYSO4kJ7SUujCH4GlfbHuke2Byo6nYrWQYAxshPM4N5IRDv51x9AHkMVaqw3vzXUSJRwQmluriRnnx4ymvjUs5w1Risr6XncMxhgvXF7Xfto3OqVCkv1oPPGvBrSCswxCFulQ+WQ1Hq1iQzOA/iXpeBJHd6s3Jr1NptPwrZWGkySOkzcmnabRzd0tSmpuuhrWQmZUnlS4NS3nnOZKwyUrSrgr0UD1CaVEpRsZI4WoiSmWA2KI9MzaXPiqNGvrpFJTcxltMLSLcSXCjDVyrgTokQfdVyLs4EFiXUPP14Zy/ggk3/PjpI001S737PbxOYOdw4PFdmjqme6alsZ2INCQzXIPaOM/ovUa9izUaqszqE3/x1Yo+iBpznamv8uP8TH2Uc8+HrJrO2kavr/hrKgranO/OuByTg44rxSSvehosEtttIFK33umUrHN0PJTkuaJ7lS32wW7/CrN0TvU/PEdap68Q83xv6Y56Fb9YGs52ViOVg/OAgv/GxxleK/CW28ZV6zueiXLMlpf7C+aXpG1Piud8OvxGc3Jlquno3Ph9+2vNGPbKjmOeoCZ6Ub17S+wIY/i43lEx2J1Rnc0S7uu3mFPX3jmAsC5pz/DXHowjPW5PeDD4mAKMIxFYXH+p3xmaD7Wh2mbOT0zFDNDMRbl8qTmg8VxYxJ9uTNNkiiKY2xG7XHgQkGKzVscw9fNhmkDBBYHIv3eXONPG6+Q1+sAe6avVQiWKV6JWKb4XIPHPW+ASBL308biAAJ7CljtQHx3HKgpNyaKDodMlzZsBeOeJME8UIvuGtW/NdyzE8PH/XywVRJFSeL2gM+tIIowD6xG3IMpAA2YJ4rMe/DsfRQc3lNB/wfC8hdQSwMEFAAAAAgAJHikXJeKuxzAAAAAEwIAAAsAAABfcmVscy8ucmVsc52SuW7DMAxAf8XQnjAH0CGIM2XxFgT5AVaiD9gSBYpFnb+v2qVxkAsZeT08EtweaUDtOKS2i6kY/RBSaVrVuAFItiWPac6RQq7ULB41h9JARNtjQ7BaLD5ALhlmt71kFqdzpFeIXNedpT3bL09Bb4CvOkxxQmlISzMO8M3SfzL38ww1ReVKI5VbGnjT5f524EnRoSJYFppFydOiHaV/Hcf2kNPpr2MitHpb6PlxaFQKjtxjJYxxYrT+NYLJD+x+AFBLAwQUAAAACAAkeKRcvaH495IBAACIBAAADwAAAHhsL3dvcmtib29rLnhtbLWUQW+bQBCF/8pq7w2OY0eqFSJFjttGalorrrgvMIRRlh06O8Rtfn0HECpSKysXn2DerB7fPO1wcyR+yYlezK/Gh5jaWqTdJEksamhcvKAWgnYq4saJlvycxJbBlbEGkMYny8XiOmkcBnt7M3ntOZkXJFAIUlCxFzKEY/zb70vzihFz9Ci/Uzu8e7CmwYANvkGZ2oU1sabjF2J8oyDOHwom71N7OTYyYMHiH/nQQ/5weRwUcfmTU5DUXi/UsEKOMpwY/J0yvoIeHqtO6BN6Ab53Ap+ZuhbDc2+jUySzMYYcpucY4obfEyNVFRZwT0XXQJAxRwbfA4ZYYxutCa6B1D65tiUWs8se+6n0Mw/lOKEo2iwv3qA2+KEcIM8HtM/MB7P3LgQoTeZ8BzOu5Qmu5Xm5dj3XzvH/sK5OYF2dF+tuq1h3hXTOmy1FmVGtTlCtzku1pY5zMAej/o8gjD87iDO09Qm09bAE080voUKN/JvaRtV1C4s9m/4x3NPlan35Ubet836r2vfwlVw5LdL0E7j9A1BLAwQUAAAACAAkeKRchTlInccAAAA8BAAAGgAAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzxZRNDoIwEEavQnoARgExMcDKDVvjBZo6UMJPm84Y8faiLKCJCzeGVfNN0/e9zTS7YCe5MQPpxlIw9t1AudDM9gRASmMvKTQWh+mmMq6XPEVXg5WqlTVCtNul4NYMUWRrZnB9WvyFaKqqUXg26t7jwF/A8DCuJY3IIrhKVyPnAsZuGRN8jn04kUVQ3nLhyttewNZCkScUbS8Ue0Lx9kKJJ5RsL3TwhA5/FCJ+dkiLzZy9+vSP9Ty9xaX9E+ehv0bHtwN4n0XxAlBLAwQUAAAACAAkeKRcokmUz0YBAADjBgAAEwAAAFtDb250ZW50X1R5cGVzXS54bWzNlcluwkAMhl8lyhWRaelyqIBL22vLoS8wnXHIiNk0Ntvb10kAqRVNQUEqlzgT2//nJUrGH9sImG2c9TjJK6L4JASqCpzEIkTw7ClDcpL4mOYiSrWQcxCjm5tHoYIn8DSkWiOfjl+glEtL2euGH6MJfpInsJhnz21gzZrkMkZrlCT2i5XXPyjDHaHgzCYGKxNxwAG5OEqoPb8DdnnvK0jJaMhmMtGbdBwlNlYgbS1g0S1xpMZQlkaBDmrpOKXAmEBqrADI2aIVHXSTiScM7fW2N7+R6QJy5CyFiLyxBOfj9iups4eRhSCR6W7xQGTp3v1BvW0N+kQ2j3cd0qLZB4rG9J/x9x0f9M+sY3QlddxdSR33V1LHwz/VoZNcGz/H/U3/93Qn9AdXVXzE1lyMyX03gifM/jOExaU/u7UtnDR+zxfNv236BVBLAQIUAxQAAAAIACR4pFxGx01IlQAAAM0AAAAQAAAAAAAAAAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQDFAAAAAgAJHikXBYKI6rvAAAAKwIAABEAAAAAAAAAAAAAAIABwwAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQDFAAAAAgAJHikXJlcnCMQBgAAnCcAABMAAAAAAAAAAAAAAIAB4QEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAMUAAAACAAkeKRclZ4lDhMBAADMAQAAGAAAAAAAAAAAAAAAgIEiCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQDFAAAAAgAJHikXJWeJQ4TAQAAzAEAABgAAAAAAAAAAAAAAICBawkAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBLAQIUAxQAAAAIACR4pFyVniUOEwEAAMwBAAAYAAAAAAAAAAAAAACAgbQKAAB4bC93b3Jrc2hlZXRzL3NoZWV0My54bWxQSwECFAMUAAAACAAkeKRclZ4lDhMBAADMAQAAGAAAAAAAAAAAAAAAgIH9CwAAeGwvd29ya3NoZWV0cy9zaGVldDQueG1sUEsBAhQDFAAAAAgAJHikXAUC4WwZBAAAIhYAABgAAAAAAAAAAAAAAICBRg0AAHhsL3dvcmtzaGVldHMvc2hlZXQ1LnhtbFBLAQIUAxQAAAAIACR4pFzlp23eJwEAAMECAAAYAAAAAAAAAAAAAACAAZURAAB4bC9kcmF3aW5ncy9kcmF3aW5nMS54bWxQSwECFAMUAAAACAAkeKRcjtHpOYsAAADtAAAAIwAAAAAAAAAAAAAAgAHyEgAAeGwvZHJhd2luZ3MvX3JlbHMvZHJhd2luZzEueG1sLnJlbHNQSwECFAMUAAAACAAkeKRcd/cihY0AAADzAAAAIwAAAAAAAAAAAAAAgAG+EwAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDUueG1sLnJlbHNQSwECFAMUAAAACAAkeKRcQ7mUqgADAACOEQAAFAAAAAAAAAAAAAAAgAGMFAAAeGwvY2hhcnRzL2NoYXJ0MS54bWxQSwECFAMUAAAACAAkeKRcqMbrCyUDAABNEAAADQAAAAAAAAAAAAAAgAG+FwAAeGwvc3R5bGVzLnhtbFBLAQIUAxQAAAAIACR4pFyXirscwAAAABMCAAALAAAAAAAAAAAAAACAAQ4bAABfcmVscy8ucmVsc1BLAQIUAxQAAAAIACR4pFy9ofj3kgEAAIgEAAAPAAAAAAAAAAAAAACAAfcbAAB4bC93b3JrYm9vay54bWxQSwECFAMUAAAACAAkeKRchTlInccAAAA8BAAAGgAAAAAAAAAAAAAAgAG2HQAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECFAMUAAAACAAkeKRcokmUz0YBAADjBgAAEwAAAAAAAAAAAAAAgAG1HgAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLBQYAAAAAEQARAIAEAAAsIAAAAAA="

def build(data):
    tasks   = data.get('tasks',[])
    config  = data.get('config',{})
    metrics = data.get('metrics',{})
    N       = int(config.get('nb_periods',12))
    CP      = int(config.get('current_period',0))
    LABELS  = (config.get('period_labels') or []) + [f'P{i+1}' for i in range(N)]
    LABELS  = LABELS[:N]
    PROJECT = config.get('project_title','Projet')
    MANAGER = config.get('manager','Chef de Projet')
    BAC = float(metrics.get('bac',0) or 0)
    EAC = float(metrics.get('eac',BAC) or BAC)
    ETC = float(metrics.get('etc',0) or 0)
    VAC = float(metrics.get('vac',0) or 0)
    CPI = float(metrics.get('cpi',1) or 1)
    SPI = float(metrics.get('spi',1) or 1)
    TCPI= float(metrics.get('tcpi',1) or 1)
    CV  = float(metrics.get('cv',0) or 0)
    SV  = float(metrics.get('sv',0) or 0)
    EV  = float(metrics.get('ev',0) or 0)
    AC  = float(metrics.get('ac',0) or 0)
    PV_CUM = [float(v or 0) for v in (metrics.get('pvC') or [0]*N)]
    EV_CUM = [(float(v) if v is not None else None) for v in (metrics.get('evC') or [None]*N)]
    AC_CUM = [(float(v) if v is not None else None) for v in (metrics.get('acC') or [None]*N)]
    while len(PV_CUM)<N: PV_CUM.append(PV_CUM[-1] if PV_CUM else 0)
    while len(EV_CUM)<N: EV_CUM.append(None)
    while len(AC_CUM)<N: AC_CUM.append(None)

    thin=Side(style='thin',color='CCCCCC')
    brd=Border(left=thin,right=thin,top=thin,bottom=thin)

    def H(ws,r,c,txt,bg='1F3864',fg='FFFFFF',sz=10):
        cell=ws.cell(r,c,txt)
        cell.font=Font(bold=True,color=fg,name='Arial',size=sz)
        cell.fill=PatternFill('solid',start_color=bg)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=brd; return cell

    def D(ws,r,c,v,fmt='#,##0',col='000000',bold=False,bg='FFFFFF'):
        cell=ws.cell(r,c,v)
        cell.number_format=fmt; cell.font=Font(name='Arial',size=10,bold=bold,color=col)
        cell.alignment=Alignment(horizontal='right',vertical='center')
        cell.fill=PatternFill('solid',start_color=bg); cell.border=brd; return cell

    def L(ws,r,c,txt,bold=False,col='000000',bg='FFFFFF'):
        cell=ws.cell(r,c,txt)
        cell.font=Font(name='Arial',size=10,bold=bold,color=col)
        cell.alignment=Alignment(horizontal='left',vertical='center')
        cell.fill=PatternFill('solid',start_color=bg); cell.border=brd; return cell

    # Load template (has chart in Courbe S)
    wb=load_workbook(io.BytesIO(base64.b64decode(TEMPLATE_B64)))

    # ── RAPPORT PAGE DE GARDE ────────────────────────────
    ws0=wb['Rapport EVM']; ws0.sheet_view.showGridLines=False
    ws0.merge_cells('B2:H3'); ws0['B2']='EARNED VALUE MANAGEMENT'
    ws0['B2'].font=Font(bold=True,size=20,name='Arial',color='FFFFFF')
    ws0['B2'].fill=PatternFill('solid',start_color='1F3864')
    ws0['B2'].alignment=Alignment(horizontal='center',vertical='center')
    ws0.row_dimensions[2].height=35; ws0.row_dimensions[3].height=35
    ws0.merge_cells('B4:H5'); ws0['B4']='Rapport de suivi EVM — Methode PMBOK 7'
    ws0['B4'].font=Font(size=13,name='Arial',color='1F3864')
    ws0['B4'].alignment=Alignment(horizontal='center',vertical='center')
    ws0['B4'].fill=PatternFill('solid',start_color='E8F0FE')
    ws0.row_dimensions[4].height=30; ws0.row_dimensions[5].height=30
    info=[('Projet',PROJECT,'5B21B6'),
          ('Periode courante',LABELS[CP] if CP<len(LABELS) else f'P{CP+1}','374151'),
          ('Date rapport',config.get('report_date',''),'374151'),
          ('Prepare par',MANAGER,'374151'),
          ('BAC Total',f'{int(BAC):,} EUR','1F3864'),
          ('EAC Prevision',f'{int(EAC):,} EUR','B45309'),
          ('CPI',f'{CPI:.2f} — {"Sous budget" if CPI>=1 else "Depassement"}','059669' if CPI>=1 else 'DC2626'),
          ('SPI',f'{SPI:.2f} — {"En avance" if SPI>=1 else "En retard"}','059669' if SPI>=1 else 'DC2626'),
          ('VAC',f'{int(VAC):,} EUR','059669' if VAC>=0 else 'DC2626'),
          ('ETC',f'{int(ETC):,} EUR','7C3AED')]
    for idx,(k,v,vc) in enumerate(info,7):
        ws0.row_dimensions[idx].height=22
        ws0.cell(idx,2,k).font=Font(bold=True,name='Arial',size=11)
        ws0.cell(idx,2).fill=PatternFill('solid',start_color='F1F5F9'); ws0.cell(idx,2).border=brd
        ws0.merge_cells(f'C{idx}:H{idx}')
        ws0.cell(idx,3,v).font=Font(name='Arial',size=11,color=vc,bold=True)
        for c in range(3,9): ws0.cell(idx,c).fill=PatternFill('solid',start_color='FFFFFF'); ws0.cell(idx,c).border=brd
    ws0.column_dimensions['A'].width=3; ws0.column_dimensions['B'].width=22
    for j in range(3,9): ws0.column_dimensions[get_column_letter(j)].width=14

    # ── PV ───────────────────────────────────────────────
    ws1=wb['PV - Planned Value']; ws1.sheet_view.showGridLines=False
    ws1.merge_cells(f'A1:{get_column_letter(N+5)}1')
    ws1['A1']='PLANNED VALUE (PV) — Budget planifie cumulatif par periode'
    ws1['A1'].font=Font(bold=True,size=12,name='Arial',color='1F3864')
    ws1['A1'].fill=PatternFill('solid',start_color='E8F0FE')
    ws1['A1'].alignment=Alignment(horizontal='left',vertical='center'); ws1.row_dimensions[1].height=25
    for j,h in enumerate(['WBS','Tache','Role','TBC']+LABELS+['BAC'],1):
        H(ws1,2,j,h,bg='2563EB' if 5<=j<=N+4 else '1F3864')
    ws1.row_dimensions[2].height=28
    for i,t in enumerate(tasks,3):
        bg='F0F7FF' if i%2==0 else 'FFFFFF'
        pv_p=[(float(v) if v else 0) for v in (t.get('pv_periods') or [])[:N]]
        while len(pv_p)<N: pv_p.append(0)
        s=0; pvc_t=[]
        for v in pv_p: s+=v; pvc_t.append(round(s))
        tbc=float(t.get('tbc',0) or 0)
        L(ws1,i,1,t.get('wbs_id',''),bold=True,col='1F3864',bg=bg)
        L(ws1,i,2,t.get('name',''),bg=bg); L(ws1,i,3,t.get('role',''),bg=bg)
        D(ws1,i,4,tbc,'#,##0',col='B45309',bold=True,bg=bg)
        for j,v in enumerate(pvc_t,5):
            D(ws1,i,j,v if v>0 else None,'#,##0',col='2563EB',bg='DBEAFE' if j-5==CP else bg)
        D(ws1,i,N+5,tbc,'#,##0',col='1F3864',bold=True,bg=bg)
    tr1=len(tasks)+3
    for j in range(1,N+6): ws1.cell(tr1,j).fill=PatternFill('solid',start_color='1F3864'); ws1.cell(tr1,j).border=brd
    L(ws1,tr1,1,'TOTAL',bold=True,col='FFFFFF',bg='1F3864')
    L(ws1,tr1,2,'Cumulative PV',bold=True,col='FFFFFF',bg='1F3864')
    D(ws1,tr1,4,round(BAC),'#,##0',col='FFFFFF',bold=True,bg='1F3864')
    for j,v in enumerate(PV_CUM[:N],5):
        D(ws1,tr1,j,round(v),'#,##0',col='FFFFFF',bold=True,bg='2563EB' if j-5==CP else '1F3864')
    D(ws1,tr1,N+5,round(BAC),'#,##0',col='FFFFFF',bold=True,bg='1F3864')
    ws1.column_dimensions['A'].width=6; ws1.column_dimensions['B'].width=26
    ws1.column_dimensions['C'].width=15; ws1.column_dimensions['D'].width=10
    for j in range(5,N+6): ws1.column_dimensions[get_column_letter(j)].width=9
    ws1.freeze_panes='E3'

    # ── EV ───────────────────────────────────────────────
    ws2=wb['EV - Earned Value']; ws2.sheet_view.showGridLines=False
    ws2.merge_cells(f'A1:{get_column_letter(N+4)}1')
    ws2['A1']='EARNED VALUE (EV) — Avancement cumule et valeur acquise'
    ws2['A1'].font=Font(bold=True,size=12,name='Arial',color='065F46')
    ws2['A1'].fill=PatternFill('solid',start_color='ECFDF5')
    ws2['A1'].alignment=Alignment(horizontal='left',vertical='center'); ws2.row_dimensions[1].height=25
    for j,h in enumerate(['WBS','Tache','TBC']+[l+' %' for l in LABELS]+['EV Periode'],1):
        H(ws2,2,j,h,bg='047857' if 4<=j<=N+3 else '065F46')
    ws2.row_dimensions[2].height=35
    for i,t in enumerate(tasks,3):
        bg='F0FDF4' if i%2==0 else 'FFFFFF'
        ev_p=[(float(v) if v else 0) for v in (t.get('ev_periods') or [])[:N]]
        while len(ev_p)<N: ev_p.append(0)
        tbc=float(t.get('tbc',0) or 0)
        L(ws2,i,1,t.get('wbs_id',''),bold=True,col='065F46',bg=bg)
        L(ws2,i,2,t.get('name',''),bg=bg)
        D(ws2,i,3,tbc,'#,##0',col='B45309',bold=True,bg=bg)
        for j,p in enumerate(ev_p,4):
            c2=ws2.cell(i,j,p if p>0 else None); c2.number_format='0.0%'
            c2.font=Font(name='Arial',size=10,color='047857' if p>0 else '94A3B8',bold=p>0)
            c2.alignment=Alignment(horizontal='center',vertical='center')
            c2.fill=PatternFill('solid',start_color='DCFCE7' if j-4==CP else bg); c2.border=brd
        ev_cp=ev_p[CP] if CP<len(ev_p) else 0
        D(ws2,i,N+4,round(ev_cp*tbc),'#,##0',col='059669',bold=True,bg='DCFCE7')
    tr2=len(tasks)+3
    for j in range(1,N+5): ws2.cell(tr2,j).fill=PatternFill('solid',start_color='065F46'); ws2.cell(tr2,j).border=brd
    L(ws2,tr2,1,'TOTAL',bold=True,col='FFFFFF',bg='065F46')
    L(ws2,tr2,2,'Cumulative EV',bold=True,col='FFFFFF',bg='065F46')
    D(ws2,tr2,3,round(BAC),'#,##0',col='FFFFFF',bold=True,bg='065F46')
    for j in range(4,N+4): ws2.cell(tr2,j).fill=PatternFill('solid',start_color='065F46'); ws2.cell(tr2,j).border=brd
    D(ws2,tr2,N+4,round(EV),'#,##0',col='FFFFFF',bold=True,bg='047857')
    ws2.column_dimensions['A'].width=6; ws2.column_dimensions['B'].width=26; ws2.column_dimensions['C'].width=10
    for j in range(4,N+5): ws2.column_dimensions[get_column_letter(j)].width=8
    ws2.freeze_panes='D3'

    # ── AC ───────────────────────────────────────────────
    ws3=wb['AC - Actual Cost']; ws3.sheet_view.showGridLines=False
    ws3.merge_cells(f'A1:{get_column_letter(N+6)}1')
    ws3['A1']='ACTUAL COST (AC) — Couts reels par periode'
    ws3['A1'].font=Font(bold=True,size=12,name='Arial',color='991B1B')
    ws3['A1'].fill=PatternFill('solid',start_color='FEF2F2')
    ws3['A1'].alignment=Alignment(horizontal='left',vertical='center'); ws3.row_dimensions[1].height=25
    for j,h in enumerate(['WBS','Tache','TBC']+[l+' AC' for l in LABELS]+['Total AC','CV','CV%'],1):
        H(ws3,2,j,h,bg='B91C1C' if 4<=j<=N+3 else '991B1B')
    ws3.row_dimensions[2].height=35
    for i,t in enumerate(tasks,3):
        bg='FFF5F5' if i%2==0 else 'FFFFFF'
        ac_p=[(float(v) if v else 0) for v in (t.get('ac_periods') or [])[:N]]
        while len(ac_p)<N: ac_p.append(0)
        ev_p=[(float(v) if v else 0) for v in (t.get('ev_periods') or [])[:N]]
        while len(ev_p)<N: ev_p.append(0)
        tbc=float(t.get('tbc',0) or 0)
        total=sum(ac_p); ev_cp=ev_p[CP] if CP<len(ev_p) else 0
        ev_val=round(ev_cp*tbc); cv_val=ev_val-round(total); cv_pct=cv_val/total if total>0 else 0
        L(ws3,i,1,t.get('wbs_id',''),bold=True,col='991B1B',bg=bg)
        L(ws3,i,2,t.get('name',''),bg=bg)
        D(ws3,i,3,tbc,'#,##0',col='B45309',bold=True,bg=bg)
        for j,v in enumerate(ac_p,4):
            D(ws3,i,j,v if v>0 else None,'#,##0',col='DC2626',bg='FEE2E2' if j-4==CP else bg)
        D(ws3,i,N+4,round(total),'#,##0',col='991B1B',bold=True,bg=bg)
        D(ws3,i,N+5,cv_val,'#,##0',col='059669' if cv_val>=0 else 'DC2626',bold=True,bg=bg)
        c2=ws3.cell(i,N+6,cv_pct); c2.number_format='0.0%'
        c2.font=Font(name='Arial',size=10,bold=True,color='059669' if cv_pct>=0 else 'DC2626')
        c2.fill=PatternFill('solid',start_color=bg); c2.border=brd
    tr3=len(tasks)+3
    for j in range(1,N+7): ws3.cell(tr3,j).fill=PatternFill('solid',start_color='991B1B'); ws3.cell(tr3,j).border=brd
    L(ws3,tr3,1,'TOTAL',bold=True,col='FFFFFF',bg='991B1B')
    L(ws3,tr3,2,'Cumulative AC',bold=True,col='FFFFFF',bg='991B1B')
    D(ws3,tr3,3,round(BAC),'#,##0',col='FFFFFF',bold=True,bg='991B1B')
    for j,v in enumerate(AC_CUM[:N],4):
        D(ws3,tr3,j,round(v) if v is not None else None,'#,##0',col='FFFFFF',bold=True,bg='B91C1C' if j-4==CP else '991B1B')
    D(ws3,tr3,N+4,round(AC),'#,##0',col='FFFFFF',bold=True,bg='991B1B')
    D(ws3,tr3,N+5,round(CV),'#,##0',col='FFFFFF',bold=True,bg='991B1B')
    ws3.column_dimensions['A'].width=6; ws3.column_dimensions['B'].width=26; ws3.column_dimensions['C'].width=10
    for j in range(4,N+7): ws3.column_dimensions[get_column_letter(j)].width=9
    ws3.freeze_panes='D3'

    # ── COURBE S (template has chart, just fill data) ────
    ws4=wb['Courbe S et Metriques']
    EAC_PROJ=[None]*N
    pv_cp=PV_CUM[CP] if CP<len(PV_CUM) else 0
    for i in range(CP,N):
        EAC_PROJ.append(None)
        if i<len(PV_CUM): EAC_PROJ[i]=round(AC+(PV_CUM[i]-pv_cp)/max(0.01,CPI))
    EAC_PROJ=EAC_PROJ[:N]

    for i,(lbl,pv_v,ev_v,ac_v,eac_p) in enumerate(zip(LABELS,PV_CUM,EV_CUM,AC_CUM,EAC_PROJ),3):
        bg='F5F3FF' if i%2==1 else 'FFFFFF'
        L(ws4,i,1,lbl,bold=i-3==CP,col='5B21B6',bg=bg)
        D(ws4,i,2,round(pv_v) if pv_v else 0,'#,##0',col='2563EB',bold=i-3==CP,bg='DBEAFE' if i-3==CP else bg)
        D(ws4,i,3,round(ev_v) if ev_v is not None else None,'#,##0',col='D97706',bg='FEF3C7' if i-3==CP else bg)
        D(ws4,i,4,round(ac_v) if ac_v is not None else None,'#,##0',col='059669',bg='DCFCE7' if i-3==CP else bg)
        D(ws4,i,5,round(eac_p) if eac_p else None,'#,##0',col='7C3AED',bg=bg)
    # Clear remaining rows beyond N
    for i in range(N+3,27):
        for j in range(1,6): ws4.cell(i,j,None)

    # KPIs under chart
    kr=34; ws4.merge_cells(f'A{kr}:E{kr}')
    ws4[f'A{kr}']='INDICATEURS EVM — Resume de performance'
    ws4[f'A{kr}'].font=Font(bold=True,size=11,name='Arial',color='FFFFFF')
    ws4[f'A{kr}'].fill=PatternFill('solid',start_color='5B21B6')
    ws4[f'A{kr}'].alignment=Alignment(horizontal='center',vertical='center')
    ws4.row_dimensions[kr].height=22
    for j2,h2 in enumerate(['Code','Indicateur','Valeur','Unite','Interpretation'],1):
        c2=ws4.cell(kr+1,j2,h2); c2.font=Font(bold=True,color='FFFFFF',name='Arial',size=10)
        c2.fill=PatternFill('solid',start_color='374151'); c2.border=brd
        c2.alignment=Alignment(horizontal='center',vertical='center')
    kpis=[('BAC','Budget at Completion',round(BAC),'#,##0','EUR','Budget initial','FFFFFF'),
          ('EAC','Estimate at Completion',round(EAC),'#,##0','EUR','Prevision cout final','FEF3C7' if EAC<=BAC else 'FEE2E2'),
          ('ETC','Estimate to Complete',round(ETC),'#,##0','EUR','Reste a depenser','F5F3FF'),
          ('VAC','Variance at Completion',round(VAC),'#,##0','EUR','Economies' if VAC>=0 else 'Depassement','DCFCE7' if VAC>=0 else 'FEE2E2'),
          ('CPI','Cost Performance Index',round(CPI,2),'0.00','','Sous budget' if CPI>=1 else 'Surcout','DCFCE7' if CPI>=1 else 'FEE2E2'),
          ('SPI','Schedule Performance Index',round(SPI,2),'0.00','','En avance' if SPI>=1 else 'En retard','DCFCE7' if SPI>=1 else 'FEE2E2'),
          ('TCPI','To Complete Perf. Index',round(TCPI,2),'0.00','','Atteignable' if TCPI<=1.1 else 'Difficile','DCFCE7' if TCPI<=1.1 else 'FEF3C7'),
          ('CV','Cost Variance',round(CV),'#,##0','EUR','Economies' if CV>=0 else 'Surcout','DCFCE7' if CV>=0 else 'FEE2E2'),
          ('SV','Schedule Variance',round(SV),'#,##0','EUR','En avance' if SV>=0 else 'En retard','DCFCE7' if SV>=0 else 'FEE2E2')]
    for idx2,(code2,label,val,fmt,unit,desc,bg) in enumerate(kpis,kr+2):
        ws4.row_dimensions[idx2].height=20
        L(ws4,idx2,1,code2,bold=True,col='5B21B6',bg=bg); L(ws4,idx2,2,label,bg=bg)
        D(ws4,idx2,3,val,fmt,bg=bg,bold=True); L(ws4,idx2,4,unit,bg=bg)
        ok=any(x in desc.lower() for x in ['avance','budget','nomie','hable'])
        L(ws4,idx2,5,desc,bold=True,col='059669' if ok else 'DC2626',bg=bg)
    ws4.column_dimensions['A'].width=8; ws4.column_dimensions['B'].width=28
    ws4.column_dimensions['C'].width=14; ws4.column_dimensions['D'].width=8; ws4.column_dimensions['E'].width=28

    buf2=io.BytesIO(); wb.save(buf2); return buf2.getvalue()

if __name__=='__main__':
    if len(sys.argv)<3: print("Usage: export_evm.py input.json output.xlsx"); sys.exit(1)
    with open(sys.argv[1]) as f: data=json.load(f)
    xlsx=build(data)
    with open(sys.argv[2],'wb') as f: f.write(xlsx)
    print(f"Excel written: {len(xlsx)} bytes")
